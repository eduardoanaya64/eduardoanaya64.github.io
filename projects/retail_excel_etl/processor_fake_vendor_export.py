# process_fake_vendor_export.py
import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import Border, Font
from openpyxl.styles.numbers import FORMAT_GENERAL

def process_file(file_path: str, output_folder: str):
    print(f"\nProcessing: {os.path.basename(file_path)}")

    df_raw = pd.read_excel(file_path)
    print(f"  ✓ Loaded raw data: {len(df_raw)} rows")

    df = df_raw.copy()

    # Rename by position (like your original): B -> Unit Name, D -> Description
    column_renames = {}
    if len(df.columns) > 1:
        column_renames[df.columns[1]] = "Unit Name"      # B
    if len(df.columns) > 3:
        column_renames[df.columns[3]] = "Description"    # D

    # Delete F and J by position if present
    columns_to_delete = []
    if len(df.columns) > 5:
        columns_to_delete.append(df.columns[5])  # F
    if len(df.columns) > 9:
        columns_to_delete.append(df.columns[9])  # J

    df = df.rename(columns=column_renames)
    df = df.drop(columns=[c for c in columns_to_delete if c in df.columns], errors="ignore")

    print(f"  ✓ Renamed: {column_renames}")
    print(f"  ✓ Deleted: {columns_to_delete}")

    # Clean empty rows
    df = df.dropna(how="all").copy()

    # Type conversions + detect columns for formatting
    date_columns = []
    currency_columns = []
    qty_columns = []

    for col in df.columns:
        # Date-ish
        if "date" in col.lower() or "day" in col.lower():
            df[col] = pd.to_datetime(df[col], errors="coerce")
            date_columns.append(col)

        # Currency-ish (we generated Sales: $)
        if "sales" in col.lower() and "$" in col:
            if df[col].dtype == "object":
                df[col] = (
                    df[col].astype(str)
                    .str.replace("$", "", regex=False)
                    .str.replace(",", "", regex=False)
                )
            df[col] = pd.to_numeric(df[col], errors="coerce")
            currency_columns.append(col)

        # Qty-ish
        if "qty" in col.lower():
            df[col] = pd.to_numeric(df[col], errors="coerce")
            qty_columns.append(col)

    print(f"  ✓ After cleaning: {len(df)} rows")

    # Save with formatting + worksheet name
    base = os.path.splitext(os.path.basename(file_path))[0]
    out_path = os.path.join(output_folder, f"{base}_processed.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet_name = "Simple Green Sales Report"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]

        no_border = Border()
        normal_font = Font(bold=False)

        # Remove borders + bold everywhere
        for row in ws.iter_rows():
            for cell in row:
                cell.border = no_border
                cell.font = normal_font

        # Apply number formats
        for col_idx, col_name in enumerate(df.columns, 1):
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)

                if col_name in date_columns:
                    cell.number_format = "M/D/YYYY"
                elif col_name in currency_columns:
                    cell.number_format = "$#,##0.00"
                elif col_name in qty_columns:
                    cell.number_format = "0"
                else:
                    cell.number_format = FORMAT_GENERAL

                cell.border = no_border
                cell.font = normal_font

        # Header row not bold
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.border = no_border
            cell.font = normal_font

    print(f"  ✅ Saved: {out_path}")
    return out_path

def main():
    print("Portfolio Processor - Fake Vendor Export")
    print("=" * 60)

    input_folder = input("Input folder (fake raw exports): ").strip().strip('"')
    output_folder = input("Output folder: ").strip().strip('"')
    os.makedirs(output_folder, exist_ok=True)

    files = [f for f in os.listdir(input_folder) if f.lower().endswith(".xlsx")]
    if not files:
        print("❌ No .xlsx files found.")
        return

    print(f"\nFound {len(files)} file(s). Processing...")
    for f in files:
        process_file(os.path.join(input_folder, f), output_folder)

    print("\nDone.")

if __name__ == "__main__":

    main()

